"""
Outstanding Report — Core Logic
(UI-independent, Streamlit aur CLI dono use kar sakte hain)
"""

import csv
import io
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ─────────────────────────────────────────────
# TALLY FILE PARSE
# ─────────────────────────────────────────────

def parse_tally(file_bytes):
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_row = None
    for i, row in enumerate(rows):
        if row[0] == 'Date':
            header_row = i
            break
    if header_row is None:
        raise ValueError("Header row nahi mili! Tally file check karo.")

    transactions = []
    i = header_row + 1
    while i < len(rows):
        row = rows[i]
        date_val    = row[0]
        dr_cr       = row[1]
        particulars = row[2]
        vch_type    = row[3]
        vch_no      = row[4]
        debit       = row[5]
        credit      = row[6]

        if isinstance(date_val, datetime) and dr_cr in ('Dr', 'Cr'):
            cost_cat = ''
            if i + 2 < len(rows):
                cc = rows[i + 2]
                if cc[2] and 'Primary Cost' not in str(cc[2]) and 'Being' not in str(cc[2]):
                    cost_cat = str(cc[2])

            transactions.append({
                'Date'         : date_val.date(),
                'Dr_Cr'        : dr_cr,
                'Particulars'  : str(particulars) if particulars else '',
                'Vch_Type'     : str(vch_type)    if vch_type    else '',
                'Vch_No'       : str(vch_no)      if vch_no      else '',
                'Debit'        : float(debit)      if debit       else 0.0,
                'Credit'       : float(credit)     if credit      else 0.0,
                'Cost_Category': cost_cat,
                'Txn_Type'     : 'PAYMENT' if vch_type == 'Payment New'
                                 else 'INVOICE' if vch_type == 'Purchase'
                                 else 'TDS'     if 'Tds' in str(particulars)
                                 else 'JOURNAL',
            })
        i += 1

    return pd.DataFrame(transactions)


# ─────────────────────────────────────────────
# OUTSTANDING LOGIC
# ─────────────────────────────────────────────

def calc_outstanding(df):
    opening = 0.0
    ob_rows = df[df['Particulars'].str.contains('Opening Balance', na=False)]
    if not ob_rows.empty:
        ob = ob_rows.iloc[0]
        opening = ob['Credit'] - ob['Debit']

    txn = df[~df['Particulars'].str.contains('Opening Balance', na=False)].copy()

    invoices = txn[txn['Txn_Type'] == 'INVOICE'].copy()
    payments = txn[txn['Txn_Type'] == 'PAYMENT'].copy()
    tds_rows = txn[txn['Txn_Type'] == 'TDS'].copy()
    journals = txn[txn['Txn_Type'] == 'JOURNAL'].copy()

    total_invoiced = invoices['Credit'].sum()
    total_paid     = payments['Debit'].sum()
    total_tds      = tds_rows['Debit'].sum()
    journal_dr     = journals['Debit'].sum()
    journal_cr     = journals['Credit'].sum()

    df_sorted = df.copy().sort_values('Date').reset_index(drop=True)
    balance = 0.0
    balances = []
    for _, row in df_sorted.iterrows():
        balance += row['Credit'] - row['Debit']
        balances.append(round(balance, 2))
    df_sorted['Running_Balance'] = balances

    outstanding = opening + total_invoiced - total_paid - total_tds + journal_cr - journal_dr

    summary = {
        'opening'        : round(opening, 2),
        'total_invoiced' : round(total_invoiced, 2),
        'total_paid'     : round(total_paid, 2),
        'total_tds'      : round(total_tds, 2),
        'outstanding'    : round(outstanding, 2),
    }
    return df_sorted, invoices, payments, tds_rows, summary


# ─────────────────────────────────────────────
# PROJECT-WISE SUMMARY
# ─────────────────────────────────────────────

def calc_project_wise(df):
    txn = df[~df['Particulars'].str.contains('Opening Balance', na=False)].copy()
    txn = txn[txn['Cost_Category'].str.strip() != '']

    tds_txn   = txn[txn['Txn_Type'] == 'TDS'].copy()
    other_txn = txn[txn['Txn_Type'] != 'TDS'].copy()

    grouped = other_txn.groupby('Cost_Category').agg(
        Sum_Debit=('Debit', 'sum'),
        Sum_Credit=('Credit', 'sum'),
    ).reset_index()

    tds_grouped = tds_txn.groupby('Cost_Category').agg(
        TDS_Amount=('Debit', 'sum'),
    ).reset_index()

    grouped = grouped.merge(tds_grouped, on='Cost_Category', how='left')
    grouped['TDS_Amount'] = grouped['TDS_Amount'].fillna(0.0)
    grouped['Payable']    = grouped['Sum_Debit'] + grouped['TDS_Amount'] - grouped['Sum_Credit']
    return grouped.sort_values('Cost_Category').reset_index(drop=True)


# ─────────────────────────────────────────────
# MASTER CSV PARSE + MATCHING
# ─────────────────────────────────────────────

PROJECT_CODE_CANDIDATES = [
    'Project Code', 'Project code', 'project code', 'PROJECT CODE',
    'Project Code ', ' Project Code', 'Project_Code', 'ProjectCode',
]

COMPARE_CANDIDATES = {
    'Balance'            : ['Balance', 'balance', 'Balance ', ' Balance'],
    'Final Amount'       : ['Final Amount', 'final amount', 'Final Amount '],
    'Payment Done'       : ['Payment Done', 'payment done', 'Payment Done '],
    'Total TDS Deducted' : ['Total TDS Deducted', 'total tds deducted'],
}


def clean_num(val):
    if val is None:
        return 0.0
    s = str(val).strip().replace(',', '').replace('%', '')
    if s == '' or s.lower() == 'nan':
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def find_column(fieldnames, candidates):
    norm_map = {fn.strip().lower(): fn for fn in fieldnames}
    for cand in candidates:
        if cand.strip().lower() in norm_map:
            return norm_map[cand.strip().lower()]
    return None


def parse_master_csv(file_bytes, project_col=None):
    text = file_bytes.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    rows = list(reader)

    if project_col is None:
        project_col = find_column(fieldnames, PROJECT_CODE_CANDIDATES)
    if project_col is None or project_col not in fieldnames:
        raise ValueError(f"Project Code column nahi mila. Available: {fieldnames}")

    available_compare_cols = {}
    for label, cands in COMPARE_CANDIDATES.items():
        col = find_column(fieldnames, cands)
        if col:
            available_compare_cols[label] = col

    data = {}
    for row in rows:
        code = str(row.get(project_col, '')).strip()
        if not code:
            continue
        entry = {label: clean_num(row.get(col)) for label, col in available_compare_cols.items()}
        data[code] = entry

    return data, project_col, list(available_compare_cols.keys()), fieldnames


def build_matching_data(project_df, master_data, compare_labels):
    matched_rows  = []
    matched_codes = set()

    for _, row in project_df.iterrows():
        code          = str(row['Cost_Category']).strip()
        tally_payable = row['Payable']
        master_row    = master_data.get(code)

        if master_row is not None:
            matched_codes.add(code)
            csv_vals = {label: master_row.get(label, 0.0) for label in compare_labels}
        else:
            csv_vals = {label: None for label in compare_labels}

        primary_val = None
        for pref in ['Balance', 'Final Amount']:
            if pref in csv_vals and csv_vals[pref] is not None:
                primary_val = csv_vals[pref]
                break
        if primary_val is None:
            for label in compare_labels:
                if csv_vals.get(label) is not None:
                    primary_val = csv_vals[label]
                    break

        diff = round(tally_payable - primary_val, 2) if primary_val is not None else None

        matched_rows.append({
            'Project Code': code,
            'Sum_Debit'   : row['Sum_Debit'],
            'TDS_Amount'  : row['TDS_Amount'],
            'Sum_Credit'  : row['Sum_Credit'],
            'Payable'     : tally_payable,
            'csv_vals'    : csv_vals,
            'Difference'  : diff,
            'Found_In_CSV': code in matched_codes,
        })

    tally_codes   = set(str(c).strip() for c in project_df['Cost_Category'])
    unmatched_csv = []
    for code, vals in master_data.items():
        if code not in tally_codes:
            unmatched_csv.append({
                'Project Code': code,
                'Sum_Debit'   : None, 'TDS_Amount': None,
                'Sum_Credit'  : None, 'Payable'   : None,
                'csv_vals'    : {label: vals.get(label, 0.0) for label in compare_labels},
                'Difference'  : None, 'Found_In_CSV': True,
            })

    return matched_rows, unmatched_csv


# ─────────────────────────────────────────────
# EXCEL STYLE HELPERS
# ─────────────────────────────────────────────

def _border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg="1F4E79", fg="FFFFFF", size=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, color=fg, name="Arial", size=size)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border    = _border()
    return c

def data_cell(ws, row, col, value, bg="FFFFFF", bold=False, align='left', num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=bold, name="Arial", size=9)
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border    = _border()
    if num_fmt:
        c.number_format = num_fmt
    return c

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_row(ws, text, cols, row=1, bg="1F4E79"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 28


# ─────────────────────────────────────────────
# SHEET BUILDERS
# ─────────────────────────────────────────────

def build_summary_sheet(wb, party_name, summary, file_period):
    ws = wb.create_sheet("SUMMARY")
    title_row(ws, f"OUTSTANDING REPORT — {party_name}", 4)
    ws.merge_cells('A2:D2')
    ws['A2'] = f"Period: {file_period}   |   Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws['A2'].font      = Font(italic=True, size=9, name="Arial", color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    rows_data = [
        ("Opening Balance (Carried Forward)", summary['opening'],        "DCE6F1", False),
        ("Total Invoices Raised (Credit)",     summary['total_invoiced'], "E2EFDA", False),
        ("Total Payments Received (Debit)",    summary['total_paid'],     "E2EFDA", False),
        ("TDS Deducted by Party",              summary['total_tds'],      "FFF2CC", False),
        (None, None, None, None),
        ("NET OUTSTANDING (Receivable)",       summary['outstanding'],    "C6EFCE" if summary['outstanding'] <= 0 else "FFC7CE", True),
    ]
    for r, item in enumerate(rows_data, 4):
        if item[0] is None:
            continue
        label, val, bg, bold = item
        data_cell(ws, r, 1, label, bg=bg, bold=bold, align='left')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        data_cell(ws, r, 4, val, bg=bg, bold=bold, align='right', num_fmt='#,##0.00')
        ws.row_dimensions[r].height = 20

    ws.merge_cells('A11:D12')
    c = ws['A11']
    if summary['outstanding'] <= 0:
        c.value = "✅ CLEAR — Party ka koi outstanding nahi hai"
        c.fill  = PatternFill("solid", fgColor="C6EFCE")
        c.font  = Font(bold=True, size=12, name="Arial", color="375623")
    else:
        c.value = f"⚠️  OUTSTANDING: ₹{summary['outstanding']:,.2f} RECEIVABLE FROM PARTY"
        c.fill  = PatternFill("solid", fgColor="FFC7CE")
        c.font  = Font(bold=True, size=12, name="Arial", color="9C0006")
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[11].height = 30
    ws.row_dimensions[12].height = 30
    set_col_widths(ws, [40, 5, 5, 18])
    ws.sheet_view.showGridLines = False


def build_ledger_sheet(wb, df_sorted, party_name):
    ws   = wb.create_sheet("LEDGER")
    cols = ['Date', 'Vch Type', 'Vch No.', 'Particulars', 'Cost Category',
            'Debit (₹)', 'Credit (₹)', 'Running Balance (₹)']
    title_row(ws, f"LEDGER — {party_name}", len(cols))
    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 2, c, h)
    ws.row_dimensions[2].height = 22

    for r_idx, (_, row) in enumerate(df_sorted.iterrows(), 3):
        txn_type = row.get('Txn_Type', '')
        bg = ("E2EFDA" if txn_type == 'INVOICE' else
              "FFF2CC" if txn_type == 'TDS'     else
              "DEEAF1" if txn_type == 'PAYMENT' else
              "EBF3FA" if r_idx % 2 == 0 else "FFFFFF")

        data_cell(ws, r_idx, 1, row['Date'],         bg=bg, align='center')
        data_cell(ws, r_idx, 2, row['Vch_Type'],      bg=bg)
        data_cell(ws, r_idx, 3, row['Vch_No'],        bg=bg)
        data_cell(ws, r_idx, 4, row['Particulars'],   bg=bg)
        data_cell(ws, r_idx, 5, row['Cost_Category'], bg=bg)
        data_cell(ws, r_idx, 6, row['Debit']  or '', bg=bg, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 7, row['Credit'] or '', bg=bg, align='right', num_fmt='#,##0.00')
        bal    = row['Running_Balance']
        bal_bg = "C6EFCE" if bal <= 0 else "FFC7CE"
        data_cell(ws, r_idx, 8, bal, bg=bal_bg, bold=True, align='right', num_fmt='#,##0.00')

    last = 2 + len(df_sorted)
    hdr_cell(ws, last + 1, 1, "TOTAL", bg="1F4E79")
    ws.merge_cells(start_row=last+1, start_column=1, end_row=last+1, end_column=5)
    for col_n, formula in [(6, f'=SUM(F3:F{last})'), (7, f'=SUM(G3:G{last})')]:
        data_cell(ws, last+1, col_n, formula, bg="1F4E79", bold=True, align='right', num_fmt='#,##0.00')
        ws.cell(row=last+1, column=col_n).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    set_col_widths(ws, [13, 14, 20, 38, 32, 14, 14, 20])
    ws.freeze_panes = "A3"


def build_invoice_sheet(wb, invoices, payments, tds_rows):
    ws = wb.create_sheet("INVOICE vs PAYMENT")
    title_row(ws, "INVOICE vs PAYMENT DETAIL", 4)

    hdr_cell(ws, 2, 1, "INVOICES RAISED", bg="375623", size=10)
    ws.merge_cells('A2:D2')
    for c, h in enumerate(['Date', 'Vch No.', 'Cost Category', 'Amount (₹)'], 1):
        hdr_cell(ws, 3, c, h, bg="4F8A10")
    ws.row_dimensions[3].height = 20

    inv_start = 4
    for r_idx, (_, row) in enumerate(invoices.iterrows(), inv_start):
        data_cell(ws, r_idx, 1, row['Date'],          align='center')
        data_cell(ws, r_idx, 2, row['Vch_No'])
        data_cell(ws, r_idx, 3, row['Cost_Category'])
        data_cell(ws, r_idx, 4, row['Credit'],        align='right', num_fmt='#,##0.00')

    inv_end   = inv_start + len(invoices) - 1
    pay_start = inv_end + 3
    hdr_cell(ws, pay_start - 1, 1, "PAYMENTS RECEIVED", bg="1F4E79", size=10)
    ws.merge_cells(start_row=pay_start-1, start_column=1, end_row=pay_start-1, end_column=4)
    for c, h in enumerate(['Date', 'Vch No.', 'Cost Category', 'Amount (₹)'], 1):
        hdr_cell(ws, pay_start, c, h, bg="2E75B6")
    for r_idx, (_, row) in enumerate(payments.iterrows(), pay_start + 1):
        data_cell(ws, r_idx, 1, row['Date'],          align='center')
        data_cell(ws, r_idx, 2, row['Vch_No'])
        data_cell(ws, r_idx, 3, row['Cost_Category'], bg="DEEAF1")
        data_cell(ws, r_idx, 4, row['Debit'],         bg="DEEAF1", align='right', num_fmt='#,##0.00')

    if not tds_rows.empty:
        tds_start = pay_start + len(payments) + 3
        hdr_cell(ws, tds_start - 1, 1, "TDS DEDUCTED", bg="974706", size=10)
        ws.merge_cells(start_row=tds_start-1, start_column=1, end_row=tds_start-1, end_column=4)
        for c, h in enumerate(['Date', 'Vch No.', 'Cost Category', 'TDS Amount (₹)'], 1):
            hdr_cell(ws, tds_start, c, h, bg="C55A11")
        for r_idx, (_, row) in enumerate(tds_rows.iterrows(), tds_start + 1):
            data_cell(ws, r_idx, 1, row['Date'],          align='center')
            data_cell(ws, r_idx, 2, row['Vch_No'])
            data_cell(ws, r_idx, 3, row['Cost_Category'], bg="FFF2CC")
            data_cell(ws, r_idx, 4, row['Debit'],         bg="FFF2CC", align='right', num_fmt='#,##0.00')

    set_col_widths(ws, [13, 22, 36, 16])
    ws.freeze_panes = "A2"


def build_project_sheet(wb, df, party_name):
    ws         = wb.create_sheet("PROJECT WISE")
    project_df = calc_project_wise(df)
    cols       = ['Project Code', 'Sum of Debit (₹)', 'TDS Amount (₹)', 'Sum of Credit (₹)', 'Payable (₹)', 'Status']
    title_row(ws, f"PROJECT WISE SUMMARY — {party_name}", len(cols))

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ws['A2'] = f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}   |   Total Projects: {len(project_df)}"
    ws['A2'].font      = Font(italic=True, size=9, name="Arial", color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 3, c, h)
    ws.row_dimensions[3].height = 22

    totals = [0.0, 0.0, 0.0, 0.0]
    for r_idx, (_, row) in enumerate(project_df.iterrows(), 4):
        payable = row['Payable']
        totals[0] += row['Sum_Debit']
        totals[1] += row['TDS_Amount']
        totals[2] += row['Sum_Credit']
        totals[3] += payable

        if abs(payable) < 1:
            row_bg, status = "C6EFCE", "✅ SETTLED"
        elif payable > 0:
            row_bg, status = "FFC7CE", "⚠️ RECEIVABLE"
        else:
            row_bg, status = "DEEAF1", "💙 EXCESS PAID"

        data_cell(ws, r_idx, 1, row['Cost_Category'], bg=row_bg)
        data_cell(ws, r_idx, 2, row['Sum_Debit'],  bg=row_bg, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 3, row['TDS_Amount'], bg="FFF2CC", align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 4, row['Sum_Credit'], bg=row_bg, align='right', num_fmt='#,##0.00')
        pay_c = data_cell(ws, r_idx, 5, payable, bg=row_bg, bold=True, align='right', num_fmt='#,##0.00')
        if payable > 1:
            pay_c.font = Font(bold=True, color="9C0006", name="Arial", size=9)
        elif payable < -1:
            pay_c.font = Font(bold=True, color="1F4E79", name="Arial", size=9)
        data_cell(ws, r_idx, 6, status, bg=row_bg)
        ws.row_dimensions[r_idx].height = 18

    tr = 4 + len(project_df)
    hdr_cell(ws, tr, 1, "GRAND TOTAL", bg="1F4E79")
    for col_n, val in zip([2, 3, 4], totals[:3]):
        data_cell(ws, tr, col_n, round(val, 2), bg="1F4E79", bold=True, align='right', num_fmt='#,##0.00')
        ws.cell(row=tr, column=col_n).font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    pay_bg = "FFC7CE" if totals[3] > 1 else "C6EFCE"
    data_cell(ws, tr, 5, round(totals[3], 2), bg=pay_bg, bold=True, align='right', num_fmt='#,##0.00')
    ws.cell(row=tr, column=5).font = Font(bold=True, color="9C0006" if totals[3] > 1 else "375623", name="Arial", size=10)
    data_cell(ws, tr, 6, "", bg=pay_bg)
    ws.row_dimensions[tr].height = 22

    lr = tr + 2
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=len(cols))
    ws.cell(row=lr, column=1).value = (
        "Legend:  ✅ SETTLED = (Debit+TDS)≈Credit   |   "
        "⚠️ RECEIVABLE = (Debit+TDS)>Credit   |   💙 EXCESS PAID = Credit>(Debit+TDS)"
    )
    ws.cell(row=lr, column=1).font      = Font(italic=True, size=8, name="Arial", color="555555")
    ws.cell(row=lr, column=1).alignment = Alignment(horizontal='left', vertical='center')

    set_col_widths(ws, [42, 16, 14, 18, 16, 16])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    return project_df   # return karo taaki main me use ho sake


def build_matching_sheet(wb, matched_rows, unmatched_csv, compare_labels, party_name, csv_name):
    ws        = wb.create_sheet("MATCHING")
    base_cols = ['Project Code', 'Sum of Debit (₹)', 'TDS Amount (₹)', 'Sum of Credit (₹)', 'Tally Payable (₹)']
    csv_cols  = [f"CSV {label} (₹)" for label in compare_labels]
    cols      = base_cols + csv_cols + ['Difference (₹)', 'Match Status']

    title_row(ws, f"TALLY vs MASTER CSV MATCHING — {party_name}", len(cols))
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(cols))
    ws['A2'] = (f"Master CSV: {csv_name}   |   "
                f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}   |   "
                f"Total Rows: {len(matched_rows) + len(unmatched_csv)}")
    ws['A2'].font      = Font(italic=True, size=9, name="Arial", color="555555")
    ws['A2'].alignment = Alignment(horizontal='center')

    for c, h in enumerate(cols, 1):
        hdr_cell(ws, 3, c, h)
    ws.row_dimensions[3].height = 24

    r_idx = 4
    counts = {'matched': 0, 'mismatch': 0, 'not_csv': 0, 'not_tally': 0}

    for row in matched_rows:
        if not row['Found_In_CSV']:
            row_bg, status = "FFF2CC", "❌ NOT IN CSV"
            counts['not_csv'] += 1
        elif row['Difference'] is not None and abs(row['Difference']) < 1:
            row_bg, status = "C6EFCE", "✅ MATCHED"
            counts['matched'] += 1
        elif row['Difference'] is not None:
            row_bg = "FFC7CE"
            status = f"⚠️ MISMATCH (₹{row['Difference']:,.0f})"
            counts['mismatch'] += 1
        else:
            row_bg, status = "DEEAF1", "ℹ️ NO COMPARE COLUMN"

        data_cell(ws, r_idx, 1, row['Project Code'], bg=row_bg)
        data_cell(ws, r_idx, 2, row['Sum_Debit'],  bg=row_bg, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 3, row['TDS_Amount'], bg=row_bg, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 4, row['Sum_Credit'], bg=row_bg, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, 5, row['Payable'],    bg=row_bg, bold=True, align='right', num_fmt='#,##0.00')

        col_n = 6
        for label in compare_labels:
            val = row['csv_vals'].get(label)
            data_cell(ws, r_idx, col_n, val if val is not None else '', bg=row_bg, align='right', num_fmt='#,##0.00')
            col_n += 1

        diff = row['Difference']
        data_cell(ws, r_idx, col_n,   diff if diff is not None else '', bg=row_bg, bold=True, align='right', num_fmt='#,##0.00')
        data_cell(ws, r_idx, col_n+1, status, bg=row_bg)
        ws.row_dimensions[r_idx].height = 18
        r_idx += 1

    for row in unmatched_csv:
        row_bg, status = "DEEAF1", "❌ NOT IN TALLY"
        counts['not_tally'] += 1
        data_cell(ws, r_idx, 1, row['Project Code'], bg=row_bg)
        for c in range(2, 6):
            data_cell(ws, r_idx, c, '', bg=row_bg, align='right')
        col_n = 6
        for label in compare_labels:
            val = row['csv_vals'].get(label)
            data_cell(ws, r_idx, col_n, val if val is not None else '', bg=row_bg, align='right', num_fmt='#,##0.00')
            col_n += 1
        data_cell(ws, r_idx, col_n,   '', bg=row_bg, align='right')
        data_cell(ws, r_idx, col_n+1, status, bg=row_bg)
        ws.row_dimensions[r_idx].height = 18
        r_idx += 1

    sr = r_idx + 1
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=len(cols))
    ws.cell(row=sr, column=1).value = (
        f"Summary:  ✅ Matched={counts['matched']}   ⚠️ Mismatch={counts['mismatch']}   "
        f"❌ Not in CSV={counts['not_csv']}   ❌ Not in Tally={counts['not_tally']}"
    )
    ws.cell(row=sr, column=1).font      = Font(bold=True, size=9, name="Arial", color="1F4E79")
    ws.cell(row=sr, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[sr].height = 20

    lr = sr + 1
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=len(cols))
    ws.cell(row=lr, column=1).value = (
        "Legend:  ✅ MATCHED≈CSV Balance/Final Amount   ⚠️ MISMATCH=Difference exists   "
        "❌ NOT IN CSV/TALLY=Project code doosri sheet me nahi"
    )
    ws.cell(row=lr, column=1).font      = Font(italic=True, size=8, name="Arial", color="555555")
    ws.cell(row=lr, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[lr].height = 18

    set_col_widths(ws, [42, 14, 12, 14, 14] + [16] * len(csv_cols) + [14, 22])
    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False

    return counts


# ─────────────────────────────────────────────
# MAIN REPORT BUILDER (returns bytes)
# ─────────────────────────────────────────────

def generate_report(tally_bytes, tally_filename,
                    csv_bytes=None, csv_filename=None,
                    csv_project_col=None):
    """
    Returns: (excel_bytes, summary_dict, match_counts_or_None, error_or_None)
    """
    try:
        df = parse_tally(tally_bytes)
    except Exception as e:
        return None, None, None, f"Tally parse error: {e}"

    df_sorted, invoices, payments, tds_rows, summary = calc_outstanding(df)
    party_name  = tally_filename.rsplit('.', 1)[0]
    min_date    = df['Date'].min()
    max_date    = df['Date'].max()
    file_period = f"{min_date.strftime('%d-%b-%Y')} to {max_date.strftime('%d-%b-%Y')}"

    wb = Workbook()
    wb.remove(wb.active)
    build_summary_sheet(wb, party_name, summary, file_period)
    build_ledger_sheet(wb, df_sorted, party_name)
    build_invoice_sheet(wb, invoices, payments, tds_rows)
    project_df = build_project_sheet(wb, df, party_name)

    match_counts = None

    if csv_bytes is not None:
        try:
            master_data, proj_col, compare_labels, fieldnames = parse_master_csv(
                csv_bytes, project_col=csv_project_col
            )
            matched_rows, unmatched_csv = build_matching_data(project_df, master_data, compare_labels)
            match_counts = build_matching_sheet(
                wb, matched_rows, unmatched_csv, compare_labels,
                party_name, csv_filename or "master.csv"
            )
        except ValueError as e:
            # CSV mila but column nahi — caller ko bata do
            return None, summary, None, f"CSV_COL_ERROR|{str(e)}"
        except Exception as e:
            return None, summary, None, f"CSV error: {e}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), summary, match_counts, None
